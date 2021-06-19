from openpyxl import load_workbook
from prettytable import PrettyTable


DATA_FILE, DATA_SHEET_NAME = 'data.xlsx', 'prices'


def print_logo():
    logo = '''

        ██████╗ ███████╗██╗  ██╗ █████╗ 
        ██╔══██╗██╔════╝██║ ██╔╝██╔══██╗
        ██║  ██║█████╗  █████╔╝ ███████║
        ██║  ██║██╔══╝  ██╔═██╗ ██╔══██║
        ██████╔╝███████╗██║  ██╗██║  ██║
        ╚═════╝ ╚══════╝╚═╝  ╚═╝╚═╝  ╚═╝

    '''
    print(logo)


def user_input():
    user_file = input('Enter file name (Press "Enter" for default "custom.xlsx"): ').strip()
    if user_file == '': user_file = 'custom.xlsx'
    user_sheet_name = input('Enter sheet name (Press "Enter" for default "Лист1"): ').strip()
    if user_sheet_name == '': user_sheet_name = 'Лист1'
    price_column = input('Enter price column (Press "Enter" for default "P"): ').strip().upper()
    if price_column == '': price_column = 'P'
    return user_file, user_sheet_name, price_column


def print_user_input(user_file, user_sheet_name, price_column):
    print()
    pretty_table = PrettyTable()
    pretty_table.field_names = ['Parameter', 'Value']
    pretty_table.add_rows([['Your file', user_file],
                           ['Target sheet', user_sheet_name],
                           ['Price column', price_column]])
    print(pretty_table)


def insert_prices(data_file, data_sheet_name, user_file, user_sheet_name, price_column):
    data_book = load_workbook(filename=data_file)
    data_sheet = data_book[data_sheet_name]
    user_book = load_workbook(filename=user_file)
    user_sheet = user_book[user_sheet_name]

    summary = PrettyTable()
    summary.field_names = ['Code', 'Price']

    for user_row in user_sheet:
        if user_row[0].value is None: break
        if user_row[0].value in ('Код', 'Code'): continue

        final_price = 'Didn\'t find'
        for data_row in data_sheet:
            if data_row[0].value == user_row[0].value:
                price = data_row[1].value
                final_price = round(price * 1.04, 2)
                user_sheet[price_column + str(user_row[0].row)].value = final_price
                break

        summary.add_row([user_row[0].value, final_price])

    user_book.save(filename=user_file)
    print('Summary:'.center(28))
    print(summary)


def run():
    print_logo()
    user_items = user_input()
    print_user_input(*user_items)
    print('\nRun...\n')
    try:
        insert_prices(DATA_FILE, DATA_SHEET_NAME, *user_items)
    except Exception as e:
        print('Error!\n', e, '\n')
    else:
        print('Done!')
    input('''(:`--..___...-''``-._             |`._
  ```--...--.      . `-..__      .`/ _\  
            `\     '       ```--`.    />
            : :   :               `:`-'
             `.:.  `.._--...___     ``--...__      
                ``--..,)       ```----....__,) Press "Enter" to exit.''')


if __name__ == '__main__':
    run()
